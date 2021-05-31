import time

import dload
import openpyxl
from selenium import webdriver
from bs4 import BeautifulSoup
from selenium.webdriver.common.keys import Keys

driver = webdriver.Chrome('./chromedriver')
html = driver.page_source
soup = BeautifulSoup(html,'html.parser',from_encoding='utf-8')
driver.set_window_position(0,0)
driver.set_window_size(10,10)

# url 카테고리, 61번 라인에 key 값으로 활용
cate = {'cpu':'873','memory':'874','main':'875','gpu':'876',"hdd":'877','case':'879','power':'880','cool':'887','ssd':'32617'}


list_all = []
price = []
title = []
imgs = []

def f_get_list(item, page):
    url = 'http://shop.danawa.com/virtualestimate/?controller=estimateMain&methods=product&categorySeq='
    url += item
    url += '''&categoryDepth=2&marketPlaceSeq=16&sellerSeq=0&pseq=2&name=&listPerPage=30
    &attribute=&makerCode=&brandCode=&serviceSectionSeq=0&productRegisterAreaGroupSeq=0
    &ignoreKeywordYN=N&preparationSale=&tabOrderbyYn=N&suggestName=&displayOptionCount=3
    &categoryName=%EA%B7%B8%EB%9E%98%ED%94%BD%EC%B9%B4%EB%93%9C&productRegisterAreaSeq=4
    &categorySeq1=861&categorySeq2=876&categorySeq3=0&categorySeq4=0&selectOptionList[]=654
    &selectOptionList[]=655&selectOptionList[]=6857&selectOptionList[]=658&selectOptionList[]=657
    &selectOptionList[]=659&selectOptionList[]=6858&selectOptionList[]=665&selectOptionList[]=661
    &selectOptionList[]=663&selectOptionList[]=664&selectOptionList[]=6201&selectOptionList[]=6574
    &selectOptionList[]=680&selectOptionList[]=16333&selectOptionList[]=20550&selectOptionList[]=31689
    &selectOptionList[]=7599&selectOptionList[]=321985&selectOptionList[]=684&selectOptionList[]=666
    &selectOptionList[]=682&selectOptionList[]=32778&selectOptionList[]=37388&selectOptionList[]=32182&selectOptionList[]=321958&goodsCount=916&page='''
    url += str(page)
    driver.get(url)
    aa = driver.page_source
    soup = BeautifulSoup(aa, 'html.parser')
    time.sleep(3)

    list = driver.find_elements_by_xpath('/html/body/div/div[3]/div[2]/div[2]/table/tbody/tr')
    img = []
    table = soup.find('div',class_='scroll_box')
    tr = table.find_all('tr')
    for i in tr:
        try:
            im = 'http:'+i.find('img')['src']
            img.append(im)
        except Exception as e:
            print(e)

    # images
    for i in img:
        src = i
        imgs.append(src)

    # title, price
    for i in list:
        t = i.find_element_by_class_name('subject').text
        p = i.find_element_by_class_name('low_price').text
        title.append(t)
        price.append(p)


# 실행 : 돌릴 페이지 수
searchPage = 157
gory = 'cool'
for i in range(1,searchPage):
    f_get_list(cate[gory],i)
    print(str(i)+'번째 페이지 완료')
    print(str(searchPage-i) +'번 남음')


# 이미지 저장
j=1
for i in imgs:
    dload.save(i,f'../images/{j}.jpg')
    j+=1

wb = openpyxl.Workbook()
sheet = wb.active

g = 1
for i in title:
    sheet.cell(row= g, column=1).value = i
    g += 1
g = 1
for i in price:
    sheet.cell(row=g, column=2).value = i
    g += 1
g = 1
for i in imgs:
    sheet.cell(row=g, column=3).value = i
    g += 1
excelName = gory + '_부품데이터.xlsx'
wb.save(excelName)
driver.quit()

print('완료')
