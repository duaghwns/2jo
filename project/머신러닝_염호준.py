import dload
import openpyxl
from selenium import webdriver
from bs4 import BeautifulSoup
from selenium.webdriver.common.keys import Keys

driver = webdriver.Chrome('./chromedriver')
driver.get('http://shop.danawa.com/virtualestimate/?controller=estimateMain&methods=product&categorySeq=873&categoryDepth=2&marketPlaceSeq=16&pseq=2')
html = driver.page_source
soup = BeautifulSoup(html, 'html.parser', from_encoding='utf-8')

price = []
title = []
imgs = []

list = driver.find_elements_by_xpath('/html/body/div/div[3]/div[2]/div[2]/table/tbody/tr')
div = driver.find_element_by_class_name('scroll_box')
img = driver.find_elements_by_tag_name('img')

# images
for i in img:
    src = i.get_attribute('src')
    imgs.append(src)

# title, price
for i in list:
    t = i.find_element_by_class_name('subject').text
    p = i.find_element_by_class_name('low_price').text
    title.append(i.find_element_by_class_name('subject').text)
    price.append(i.find_element_by_class_name('low_price').text)

print(title)
print(price)
print(imgs[2:])

# 이미지 저장
j=1
for i in imgs[2:]:
    dload.save(i,f'./images/{j}.jpg')
    j+=1

wb = openpyxl.Workbook()
sheet = wb.active


g = 1
for i in title:
    sheet.cell(row= g, column=1).value = i
    g += 1
g = 1
for i in price:
    sheet.cell(row=g, column=2).value = i
    g += 1
g = 1
for i in imgs[2:]:
    sheet.cell(row=g, column=3).value = i
    g += 1

wb.save('부품데이터.xlsx')



driver.quit()

